"""
ejecutar_kqnodes.py

Ejecutor de KQNodes para las hojas del Excel de pruebas del proyecto.

Lee el Excel de pruebas (GeoMIP/tests/PruebasK-Particiones.xlsx por defecto),
ejecuta KQNodes para k en [k_min, k_max] y guarda resultados en:
    KQNodes/results/resultados_kqnodes_{hoja}{ks}.xlsx

NO modifica el Excel de entrada. El usuario copia los resultados manualmente.

KQNodes vive en un arbol separado (este, KQNodes/). El Excel de pruebas y las TPMs
grandes viven en el arbol GeoMIP; este ejecutor las resuelve por ruta relativa.
Las TPMs se buscan primero en KQNodes/src/.samples y luego en GeoMIP/data/samples.
NOTA: actualmente solo existen en el repo las TPMs N10A y N15B (las hojas 20A/22A/25A
no tienen CSV disponible y reportaran 'TPM no encontrada').

Como ejecutar (desde el directorio KQNodes/):

  Todas las k (2 a 5):
    ..\QNodes\.venv\Scripts\python.exe ejecutar_kqnodes.py --hoja 10A-Elementos

  Solo una k:
    ..\QNodes\.venv\Scripts\python.exe ejecutar_kqnodes.py --hoja 15B-Elementos --k 3

  Rango de k:
    ..\QNodes\.venv\Scripts\python.exe ejecutar_kqnodes.py --hoja 15B-Elementos --k_min 2 --k_max 4

  Con paginacion y timeout:
    ..\QNodes\.venv\Scripts\python.exe ejecutar_kqnodes.py --hoja 15B-Elementos --inicio 0 --cantidad 10 --timeout 600

Variables de entorno opcionales:
    KQNODES_INPUT_XLSX : ruta al Excel de entrada
    KQNODES_TIMEOUT_S  : timeout por prueba en segundos (default 86400 = 24h)
"""

import os
import re
import sys
import time
import logging
import argparse
import multiprocessing
from pathlib import Path

import numpy as np
import pandas as pd

# ── Rutas ─────────────────────────────────────────────────────────────────────
KQNODES_ROOT = Path(__file__).resolve().parent          # .../KQNodes
REPO_ROOT    = KQNODES_ROOT.parent                       # .../Optimal-Bipartition-Solver
sys.path.insert(0, str(KQNODES_ROOT))

# ── Silenciar logging ruidoso del arbol (SafeLogger imprime CRITICAL) ──────────
logging.disable(logging.CRITICAL)

import src.middlewares.slogger as _slogger_mod


class _NullSafeLogger:
    def __init__(self, *a, **k): pass
    def debug(self, *a, **k): pass
    def info(self, *a, **k): pass
    def warn(self, *a, **k): pass
    def error(self, *a, **k): pass
    def critic(self, *a, **k): pass
    def fatal(self, *a, **k): pass
    def set_log(self, *a, **k): pass


_slogger_mod.SafeLogger = _NullSafeLogger

from src.models.base.application import aplicacion
from src.strategies.kqnodes import KQNodes

# ── Configuracion base ─────────────────────────────────────────────────────────
TIMEOUT_S  = int(os.getenv("KQNODES_TIMEOUT_S", "86400"))  # 24h default
INPUT_XLSX = Path(os.getenv(
    "KQNODES_INPUT_XLSX",
    str(REPO_ROOT / "GeoMIP" / "tests" / "PruebasK-Particiones.xlsx"),
))


# ── Helpers ───────────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int) -> str:
    """Convierte una cadena de letras (A,B,C,...) a una mascara binaria de n_bits."""
    posiciones = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_bits]
    binario = ["0"] * n_bits
    for letra in str(texto).upper():
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def parsear_hoja(sheet_name: str) -> tuple[int, str]:
    m = re.match(r"(\d+)([A-Z])", sheet_name.strip())
    if not m:
        raise ValueError(f"No se pudo parsear '{sheet_name}'. Formato: '10A-Elementos'.")
    return int(m.group(1)), m.group(2)


def resolver_tpm(n: int, variante: str) -> tuple[Path, np.ndarray]:
    nombre = f"N{n}{variante}.csv"
    candidatos = (
        KQNODES_ROOT / "src" / ".samples" / nombre,
        KQNODES_ROOT / ".samples" / nombre,
        REPO_ROOT / "GeoMIP" / "data" / "samples" / nombre,
    )
    for p in candidatos:
        if p.exists():
            return p, np.genfromtxt(p, delimiter=",")
    raise FileNotFoundError(
        f"No se encontro '{nombre}' en: " + ", ".join(str(c) for c in candidatos)
    )


def leer_excel(ruta: Path, sheet_name: str) -> tuple[str, list[tuple]]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw = ws["B1"].value
    estado_inicial = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    filas = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and alcance_str is not None and mecanismo_str is not None:
            filas.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))
    wb.close()
    return estado_inicial, filas


# ── Worker de multiprocessing ─────────────────────────────────────────────────

def _worker(
    estado_inicial: str,
    condiciones:    str,
    alcance:        str,
    mecanismo:      str,
    tpm:            "np.ndarray",
    pagina:         str,
    k_min:          int,
    k_max:          int,
    cola:           multiprocessing.Queue,
) -> None:
    """Proceso hijo: crea KQNodes, ejecuta y pone resultados por k en la cola."""
    try:
        aplicacion.set_pagina_red_muestra(pagina)
        sia = KQNodes(tpm, k_min=k_min, k_max=k_max)
        sia.aplicar_estrategia(estado_inicial, condiciones, alcance, mecanismo)

        por_k = {}
        for k, datos in sia._resultados_por_k.items():
            por_k[k] = {
                "phi":          datos.get("phi"),
                "particion":    datos.get("particion"),
                "n_candidatos": datos.get("n_candidatos"),
                "modo_usado":   datos.get("modo_usado"),
                "tiempo_s":     datos.get("tiempo_s"),
            }
        resultado = {"por_k": por_k, "error": None}
    except Exception as exc:
        resultado = {"por_k": {}, "error": str(exc)}

    cola.put(resultado)


# ── Ejecutor principal ────────────────────────────────────────────────────────

def ejecutar_desde_excel(
    sheet_name: str,
    inicio:     int = 0,
    cantidad:   int = 50,
    k_min:      int = 2,
    k_max:      int = 5,
    timeout_s:  int = TIMEOUT_S,
    etiqueta:   str = "",
) -> None:
    n, variante = parsear_hoja(sheet_name)
    estado_inicial, todos_casos = leer_excel(INPUT_XLSX, sheet_name)
    try:
        tpm_path, tpm = resolver_tpm(n, variante)
    except FileNotFoundError as exc:
        print(f"[ABORTADO] {exc}")
        print(f"La hoja '{sheet_name}' requiere la TPM N{n}{variante}.csv, que no esta "
              "disponible en el repo. Hojas ejecutables actualmente: 10A-Elementos, 15B-Elementos.")
        return
    condiciones = "1" * len(estado_inicial)
    filas_casos = todos_casos[inicio: inicio + cantidad]

    ks_label = f"k={k_min}" if k_min == k_max else f"k={k_min}..{k_max}"

    print("=" * 72)
    print(f"KQNodes — Hoja: {sheet_name}  |  {ks_label}")
    print(f"Estado inicial: {estado_inicial}  n={len(estado_inicial)}")
    print(f"TPM: {tpm_path}")
    print(f"Pruebas: {inicio + 1} -> {inicio + len(filas_casos)}")
    print(f"Timeout por prueba: {timeout_s}s ({timeout_s/3600:.1f}h)")
    print("=" * 72)

    resultados = []

    for prueba_num, alcance_str, mecanismo_str in filas_casos:
        alcance   = convertir_a_binario(alcance_str,   len(estado_inicial))
        mecanismo = convertir_a_binario(mecanismo_str, len(estado_inicial))

        print(f"\n[{prueba_num:>3}] Alcance={alcance_str!r:<28} Mecanismo={mecanismo_str!r}")

        cola = multiprocessing.Queue()
        proceso = multiprocessing.Process(
            target=_worker,
            args=(estado_inicial, condiciones, alcance, mecanismo,
                  tpm, variante, k_min, k_max, cola),
        )
        t_ini = time.perf_counter()
        proceso.start()
        proceso.join(timeout=timeout_s)
        t_total = round(time.perf_counter() - t_ini, 4)

        if proceso.is_alive():
            print(f"  [TIMEOUT {timeout_s}s] terminado forzosamente.")
            proceso.terminate()
            proceso.join()
            por_k_data, error = {}, f"Timeout ({timeout_s}s)"
        elif cola.empty():
            por_k_data, error = {}, "Proceso termino sin resultado"
        else:
            res = cola.get()
            por_k_data = res.get("por_k", {})
            error = res.get("error")

        fila = {
            "#Prueba":   prueba_num,
            "Alcance":   alcance_str,
            "Mecanismo": mecanismo_str,
            "T_total_s": t_total,
            "Error":     error,
        }

        for k in range(2, 6):  # columnas k2-k5 para compatibilidad con el Excel
            prefix = f"k{k}"
            if k in por_k_data:
                d = por_k_data[k]
                phi_str = str(d["phi"]).replace(".", ",") if d["phi"] is not None else None
                t_str   = str(d["tiempo_s"]).replace(".", ",") if d["tiempo_s"] is not None else None
                fila[f"{prefix}_Particion"]    = d["particion"]
                fila[f"{prefix}_Perdida"]      = phi_str
                fila[f"{prefix}_Tiempo_s"]     = t_str
                fila[f"{prefix}_N_candidatos"] = d["n_candidatos"]
                fila[f"{prefix}_ModoUsado"]    = d["modo_usado"]
                if error is None and d["phi"] is not None:
                    print(f"  k={k}: phi={d['phi']:.6f}  t={d['tiempo_s']:.3f}s  "
                          f"cand={d['n_candidatos']}  modo={d['modo_usado']}")
            else:
                fila[f"{prefix}_Particion"]    = None
                fila[f"{prefix}_Perdida"]      = None
                fila[f"{prefix}_Tiempo_s"]     = None
                fila[f"{prefix}_N_candidatos"] = None
                fila[f"{prefix}_ModoUsado"]    = None

        if error:
            print(f"  ERROR: {error}")

        resultados.append(fila)

    # ── Guardar ───────────────────────────────────────────────────────────────
    df = pd.DataFrame(resultados)
    hoja_clean = sheet_name.strip().replace(" ", "_")
    ks_suffix  = f"_k{k_min}" if k_min == k_max else f"_k{k_min}-{k_max}"
    eta_suffix = f"_{etiqueta}" if etiqueta else ""
    ruta_salida = KQNODES_ROOT / "results" / f"resultados_kqnodes_{hoja_clean}{ks_suffix}{eta_suffix}.xlsx"
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(ruta_salida, index=False)

    completadas = sum(1 for r in resultados if r["Error"] is None)
    print(f"\n{'=' * 72}")
    print(f"Resultados guardados en: {ruta_salida}")
    print(f"Completadas: {completadas}/{len(resultados)}  |  Errores/Timeouts: {len(resultados) - completadas}")
    print("=" * 72)


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Ejecutor KQNodes por hojas")
    parser.add_argument("--hoja",     required=True,        help="Nombre de hoja (ej. '10A-Elementos')")
    parser.add_argument("--inicio",   type=int, default=0,  help="Indice de inicio 0-based (default 0)")
    parser.add_argument("--cantidad", type=int, default=50, help="Pruebas a ejecutar (default 50)")
    parser.add_argument("--timeout",  type=int, default=TIMEOUT_S,
                        help=f"Timeout por prueba en segundos (default {TIMEOUT_S})")
    parser.add_argument("--etiqueta", type=str, default="",
                        help="Sufijo adicional para el nombre del archivo de salida")

    k_group = parser.add_mutually_exclusive_group()
    k_group.add_argument("--k",     type=int, help="Evaluar solo esta k (ej. --k 3)")
    k_group.add_argument("--k_min", type=int, default=None, help="k minimo del rango (default 2)")
    parser.add_argument("--k_max",  type=int, default=None, help="k maximo del rango (default 5)")

    args = parser.parse_args()

    if args.k is not None:
        k_min = k_max = args.k
    else:
        k_min = args.k_min if args.k_min is not None else 2
        k_max = args.k_max if args.k_max is not None else 5

    if k_min < 2:
        parser.error("k_min debe ser >= 2")
    if k_max > 5:
        parser.error("k_max debe ser <= 5")
    if k_min > k_max:
        parser.error("k_min no puede ser mayor que k_max")

    ejecutar_desde_excel(
        sheet_name = args.hoja,
        inicio     = args.inicio,
        cantidad   = args.cantidad,
        k_min      = k_min,
        k_max      = k_max,
        timeout_s  = args.timeout,
        etiqueta   = args.etiqueta,
    )


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()