"""
experiments/prueba_determinismo_singleton.py

Verifica determinismo de KGeoMIPAsimetrico tras la separacion estricta
de candidatos_obligatorios (singletons) y candidatos_opcionales.

Corre los casos de prueba indicados dos veces y compara phi por k.
Reporta:
  (a) Si ambas corridas producen exactamente el mismo phi  -> DETERMINISTA
  (b) Cuantos casos llegan a phi=0

Como ejecutar (desde Method2_Dynamic_Programming_Reformulation/):
    python experiments/prueba_determinismo_singleton.py
"""

import sys
import re
import time
import logging
from pathlib import Path

import numpy as np
import pandas as pd

# ── Rutas ─────────────────────────────────────────────────────────────────────
METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT  = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(METHOD2_ROOT))

logging.disable(logging.CRITICAL)

import src.middlewares.slogger as _slogger_mod

class _NullSafeLogger:
    def __init__(self, name): pass
    def debug(self, *a, **k): pass
    def info(self, *a, **k): pass
    def warn(self, *a, **k): pass
    def error(self, *a, **k): pass
    def critic(self, *a, **k): pass
    def fatal(self, *a, **k): pass
    def set_log(self, *a, **k): pass

_slogger_mod.SafeLogger = _NullSafeLogger

from src.models.base.application import aplicacion
from src.middlewares.profile import profiler_manager
from src.controllers.manager import Manager
from src.controllers.strategies.kgeometric_asimetrico import KGeoMIPAsimetrico
from src.funcs.decay import decay_exponencial

profiler_manager.enabled = False

# ── Configuracion ─────────────────────────────────────────────────────────────
HOJA          = "10A-Elementos"
N             = 10
VARIANTE      = "A"
# Pruebas 1-indexadas a verificar (indices en el Excel de pruebas)
PRUEBAS_IDX   = [2, 3, 8, 16, 17, 22, 23, 25]
K_MIN         = 2
K_MAX         = 2
M_MAX_EXACTO  = 8   # modo heuristico para m > 8

INPUT_XLSX = GEOMIP_ROOT / "tests" / "PruebasK-Particiones.xlsx"


# ── Utilidades ─────────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int) -> str:
    posiciones = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto.upper():
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def leer_excel_casos(ruta: Path, sheet: str, indices: list[int]) -> tuple[str, list[tuple]]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[sheet]
    raw = ws["B1"].value
    estado_inicial = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    idx_set = set(indices)
    filas = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and int(prueba) in idx_set:
            filas.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))
    wb.close()
    filas.sort(key=lambda r: r[0])
    return estado_inicial, filas


def resolver_tpm(n: int, variante: str) -> np.ndarray:
    nombre = f"N{n}{variante}.csv"
    for base in (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT  / "data" / "samples",
    ):
        if (base / nombre).exists():
            return np.genfromtxt(base / nombre, delimiter=",")
    raise FileNotFoundError(f"No se encontro '{nombre}'")


def correr_caso(estado_inicial, condicion, alcance_bin, mecanismo_bin, tpm, variante, k_min, k_max, m_max):
    aplicacion.pagina_sample_network = variante
    gestor = Manager(estado_inicial=estado_inicial)
    sia = KGeoMIPAsimetrico(
        gestor,
        k_max=k_max,
        k_min=k_min,
        m_max_exhaustivo=m_max,
        decay_fn=decay_exponencial,
    )
    sia.aplicar_estrategia(condicion, alcance_bin, mecanismo_bin, tpm)
    return {k: datos["phi"] for k, datos in sia._resultados_por_k.items()}


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print(f"Prueba de determinismo — singleton cuts separacion estricta")
    print(f"Hoja: {HOJA}  |  k={K_MIN}  |  Pruebas: {PRUEBAS_IDX}")
    print(f"m_max_exacto={M_MAX_EXACTO}  (heuristico para m > {M_MAX_EXACTO})")
    print("=" * 70)

    tpm = resolver_tpm(N, VARIANTE)
    estado_inicial, casos = leer_excel_casos(INPUT_XLSX, HOJA, PRUEBAS_IDX)
    condicion = "1" * len(estado_inicial)

    if len(casos) != len(PRUEBAS_IDX):
        encontrados = [c[0] for c in casos]
        faltantes = [i for i in PRUEBAS_IDX if i not in encontrados]
        print(f"ADVERTENCIA: No se encontraron las pruebas {faltantes} en el Excel.")

    resultados_corrida = [[], []]

    for corrida in range(2):
        print(f"\n{'-' * 70}")
        print(f"CORRIDA {corrida + 1}/2")
        print(f"{'-' * 70}")

        for prueba_num, alcance_str, mecanismo_str in casos:
            alcance_bin   = convertir_a_binario(alcance_str,   len(estado_inicial))
            mecanismo_bin = convertir_a_binario(mecanismo_str, len(estado_inicial))

            t0 = time.perf_counter()
            try:
                phis = correr_caso(
                    estado_inicial, condicion, alcance_bin, mecanismo_bin,
                    tpm, VARIANTE, K_MIN, K_MAX, M_MAX_EXACTO
                )
                phi_k2 = phis.get(2)
                error  = None
            except Exception as exc:
                phi_k2 = None
                error  = str(exc)
            t_s = time.perf_counter() - t0

            resultados_corrida[corrida].append({
                "prueba":  prueba_num,
                "phi_k2":  phi_k2,
                "t_s":     round(t_s, 3),
                "error":   error,
            })

            estado = f"phi={phi_k2:.6f}" if phi_k2 is not None else f"ERROR: {error}"
            print(f"  [{prueba_num:>3}] {alcance_str:<30} {mecanismo_str:<30} {estado}  ({t_s:.2f}s)")

    # ── Comparacion de corridas ────────────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print("COMPARACION CORRIDA 1 vs CORRIDA 2")
    print(f"{'=' * 70}")
    print(f"{'Prueba':>8}  {'phi C1':>12}  {'phi C2':>12}  {'Igual?':>8}  {'phi=0?':>8}")
    print(f"{'-' * 8}  {'-' * 12}  {'-' * 12}  {'-' * 8}  {'-' * 8}")

    todos_iguales  = True
    count_phi_cero = 0

    for r1, r2 in zip(resultados_corrida[0], resultados_corrida[1]):
        assert r1["prueba"] == r2["prueba"]
        p1, p2 = r1["phi_k2"], r2["phi_k2"]

        if p1 is None or p2 is None:
            igual = (p1 is None and p2 is None)
        else:
            igual = abs(p1 - p2) < 1e-9

        es_cero = (p1 is not None and abs(p1) < 1e-9)

        if not igual:
            todos_iguales = False
        if es_cero:
            count_phi_cero += 1

        p1_str = f"{p1:.8f}" if p1 is not None else "None"
        p2_str = f"{p2:.8f}" if p2 is not None else "None"
        igual_str = "SI" if igual else "NO !!!"
        cero_str  = "SI" if es_cero else ""

        print(f"{r1['prueba']:>8}  {p1_str:>12}  {p2_str:>12}  {igual_str:>8}  {cero_str:>8}")

    print(f"\nDETERMINISMO: {'CONFIRMADO — ambas corridas identicas' if todos_iguales else 'FALLO — corridas divergen'}")
    print(f"Casos phi=0 : {count_phi_cero}/{len(PRUEBAS_IDX)}")
    print("=" * 70)


if __name__ == "__main__":
    main()
