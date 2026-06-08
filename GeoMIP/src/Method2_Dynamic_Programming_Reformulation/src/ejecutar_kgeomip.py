"""
src/ejecutar_kgeomip.py

Ejecutor de KGeoMIP Simetrico para las pruebas del proyecto K-QGMIP.

Lee el Excel de pruebas (GeoMIP/tests/PruebasK-Particiones.xlsx), ejecuta
KGeoMIPHeuristica y guarda resultados en:
    GeoMIP/results/resultados_kgeomip_simetrico_{hoja}.xlsx

NO modifica el Excel de entrada. El usuario copia los resultados manualmente.

Como ejecutar (desde el directorio Method2_Dynamic_Programming_Reformulation/):

  Todas las k (2 a 5):
    python src/ejecutar_kgeomip.py --hoja 10A-Elementos

  Solo una k especifica:
    python src/ejecutar_kgeomip.py --hoja 15B-Elementos --k 2
    python src/ejecutar_kgeomip.py --hoja 20A-Elementos --k 3

  Rango de k:
    python src/ejecutar_kgeomip.py --hoja 22A-Elementos --k_min 2 --k_max 3

  Con paginacion:
    python src/ejecutar_kgeomip.py --hoja "25A-Elementos " --inicio 0 --cantidad 25

Variables de entorno opcionales:
    KGEOMIP_INPUT_XLSX   : ruta al Excel de entrada
    KGEOMIP_N_MAX_EXACTO : umbral exacto/heuristico (default 8)
    KGEOMIP_TIMEOUT_S    : timeout por prueba en segundos (default 86400 = 24h)
"""

import sys
import os
import re
import time
import logging
import argparse
import multiprocessing
from pathlib import Path

import numpy as np
import pandas as pd

# ── Rutas ─────────────────────────────────────────────────────────────────────
METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT  = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(METHOD2_ROOT))

# ── Silenciar logging ─────────────────────────────────────────────────────────
logging.disable(logging.CRITICAL)

import src.middlewares.slogger as _slogger_mod

class _NullSafeLogger:
    def __init__(self, name): pass
    def debug(self, *a, **k): pass
    def info(self, *a, **k): pass
    def warn(self, *a, **k): pass
    def error(self, *a, **k): pass
    def critic(self, *a, **k): pass
    def fatal(self, *a, **k): pass
    def set_log(self, *a, **k): pass

_slogger_mod.SafeLogger = _NullSafeLogger

# ── Importaciones del proyecto ────────────────────────────────────────────────
from src.models.base.application import aplicacion
from src.middlewares.profile import profiler_manager
from src.controllers.manager import Manager
from src.controllers.strategies.kgeometric_heuristica import KGeoMIPHeuristica
from src.funcs.decay import decay_exponencial
from src.funcs.format import fmt_k_particion

profiler_manager.enabled = False

# ── Configuracion base (sobreescribible por args/env) ─────────────────────────
N_MAX_EXACTO   = int(os.getenv("KGEOMIP_N_MAX_EXACTO", "8"))
TIMEOUT_S      = int(os.getenv("KGEOMIP_TIMEOUT_S",    "86400"))  # 24h default
INPUT_XLSX     = Path(os.getenv(
    "KGEOMIP_INPUT_XLSX",
    str(GEOMIP_ROOT / "tests" / "PruebasK-Particiones.xlsx"),
))


# ── Helpers ───────────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int) -> str:
    posiciones = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto.upper():
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def parsear_hoja(sheet_name: str) -> tuple[int, str]:
    m = re.match(r"(\d+)([A-Z])", sheet_name.strip())
    if not m:
        raise ValueError(f"No se pudo parsear '{sheet_name}'. Formato: '10A-Elementos'.")
    return int(m.group(1)), m.group(2)


def resolver_tpm(n: int, variante: str) -> tuple[Path, np.ndarray]:
    nombre = f"N{n}{variante}.csv"
    for base in (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT  / "data" / "samples",
    ):
        if (base / nombre).exists():
            p = base / nombre
            return p, np.genfromtxt(p, delimiter=",")
    raise FileNotFoundError(f"No se encontro '{nombre}'")


def leer_excel(ruta: Path, sheet_name: str) -> tuple[str, list[tuple]]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw = ws["B1"].value
    estado_inicial = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    filas = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and alcance_str is not None and mecanismo_str is not None:
            filas.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))
    wb.close()
    return estado_inicial, filas


def _formatear_particion(
    asig: list,
    comunes: "np.ndarray",
) -> str:
    """
    Formatea los grupos BALANCEADOS de la particion (sin incluir huerfanos).

    Los nodos huerfanos son un detalle de implementacion para el calculo de phi
    y no deben aparecer en la columna de particion del Excel, pues inflan el
    numero de grupos mostrados (ej: k=2 con 1 huerfano mostraria 3 columnas).
    """
    import numpy as _np
    grupos_fmt = []
    for grupo in asig:
        arr = _np.array(grupo, dtype=int)
        nodos = [int(x) for x in comunes[arr]]
        grupos_fmt.append((nodos, nodos))
    return fmt_k_particion(grupos_fmt)


# ── Worker de multiprocessing ─────────────────────────────────────────────────

def _worker(
    estado_inicial: str,
    condiciones:    str,
    alcance:        str,
    mecanismo:      str,
    tpm:            "np.ndarray",
    pagina:         str,
    n_max_exacto:   int,
    k_min:          int,
    k_max:          int,
    cola:           multiprocessing.Queue,
) -> None:
    """Proceso hijo: crea KGeoMIPHeuristica, ejecuta y pone resultados en la cola."""
    try:
        aplicacion.pagina_sample_network = pagina
        gestor = Manager(estado_inicial=estado_inicial)
        sia = KGeoMIPHeuristica(
            gestor,
            k_max=k_max,
            k_min=k_min,
            n_max_exhaustivo=n_max_exacto,
            decay_fn=decay_exponencial,
        )
        sia.aplicar_estrategia(condiciones, alcance, mecanismo, tpm)

        import numpy as _np
        indices  = sia.sia_subsistema.indices_ncubos
        dims     = sia.sia_subsistema.dims_ncubos
        comunes  = _np.intersect1d(indices, dims)

        resultado = {"modo": sia._modo_usado, "por_k": {}, "error": None}

        for k, datos in sia._resultados_por_k.items():
            asig = datos["asignacion"]
            particion_str = _formatear_particion(asig, comunes) if asig else None
            resultado["por_k"][k] = {
                "phi":          datos["phi"],
                "particion":    particion_str,
                "n_candidatos": datos["n_candidatos"],
                "tiempo_s":     datos["tiempo_s"],
            }

    except Exception as exc:
        resultado = {"modo": "error", "por_k": {}, "error": str(exc)}

    cola.put(resultado)


# ── Ejecutor principal ────────────────────────────────────────────────────────

def ejecutar_desde_excel(
    sheet_name:   str,
    inicio:       int = 0,
    cantidad:     int = 50,
    k_min:        int = 2,
    k_max:        int = 5,
    timeout_s:    int = TIMEOUT_S,
    n_max_exacto: int = N_MAX_EXACTO,
    etiqueta:     str = "",
) -> None:
    n, variante = parsear_hoja(sheet_name)
    estado_inicial, todos_casos = leer_excel(INPUT_XLSX, sheet_name)
    tpm_path, tpm = resolver_tpm(n, variante)
    condiciones   = "1" * len(estado_inicial)
    filas         = todos_casos[inicio : inicio + cantidad]

    ks_label  = f"k={k_min}" if k_min == k_max else f"k={k_min}..{k_max}"
    modo_label = "EXACTO EXHAUSTIVO" if n_max_exacto >= 999 else f"n_max_exacto={n_max_exacto}"

    print("=" * 70)
    print(f"KGeoMIP Simetrico — Hoja: {sheet_name}  |  {ks_label}  |  {modo_label}")
    print(f"Estado inicial: {estado_inicial}  n={len(estado_inicial)}")
    print(f"TPM: {tpm_path}")
    print(f"Pruebas: {inicio + 1} → {inicio + len(filas)}")
    print(f"Timeout por prueba: {timeout_s}s ({timeout_s/3600:.1f}h)")
    if n_max_exacto >= 999:
        from src.funcs.partitions import contar_stirling
        stirling_max = sum(contar_stirling(n, k) for k in range(k_min, k_max + 1))
        print(f"AVISO: modo exacto — hasta {stirling_max:,} candidatos por caso (n_bal={n})")
    print("=" * 70)

    resultados = []

    for prueba_num, alcance_str, mecanismo_str in filas:
        alcance   = convertir_a_binario(alcance_str,   len(estado_inicial))
        mecanismo = convertir_a_binario(mecanismo_str, len(estado_inicial))

        print(f"\n[{prueba_num:>3}] Alcance={alcance_str!r:<30} Mecanismo={mecanismo_str!r}")

        cola    = multiprocessing.Queue()
        proceso = multiprocessing.Process(
            target=_worker,
            args=(estado_inicial, condiciones, alcance, mecanismo,
                  tpm, variante, n_max_exacto, k_min, k_max, cola),
        )
        t_ini = time.perf_counter()
        proceso.start()
        proceso.join(timeout=timeout_s)
        t_total = round(time.perf_counter() - t_ini, 4)

        if proceso.is_alive():
            print(f"  [TIMEOUT {timeout_s}s] terminado forzosamente.")
            proceso.terminate()
            proceso.join()
            por_k_data, modo, error = {}, "timeout", f"Timeout ({timeout_s}s)"
        elif cola.empty():
            por_k_data, modo, error = {}, "error", "Proceso termino sin resultado"
        else:
            res        = cola.get()
            por_k_data = res.get("por_k", {})
            modo       = res.get("modo", "?")
            error      = res.get("error")

        fila: dict = {
            "#Prueba":   prueba_num,
            "Alcance":   alcance_str,
            "Mecanismo": mecanismo_str,
            "Modo":      modo,
            "T_total_s": t_total,
            "Error":     error,
        }

        for k in range(2, 6):  # siempre columnas k2-k5 para compatibilidad con Excel
            prefix = f"k{k}"
            if k in por_k_data:
                d = por_k_data[k]
                phi_str  = str(d["phi"]).replace(".", ",") if d["phi"] is not None else None
                t_str    = str(d["tiempo_s"]).replace(".", ",") if d["tiempo_s"] is not None else None
                fila[f"{prefix}_Particion"]    = d["particion"]
                fila[f"{prefix}_Perdida"]      = phi_str
                fila[f"{prefix}_Tiempo_s"]     = t_str
                fila[f"{prefix}_N_candidatos"] = d["n_candidatos"]
                if error is None:
                    print(f"  k={k}: phi={d['phi']:.6f}  t={d['tiempo_s']:.3f}s  cand={d['n_candidatos']}")
            else:
                fila[f"{prefix}_Particion"]    = None
                fila[f"{prefix}_Perdida"]      = None
                fila[f"{prefix}_Tiempo_s"]     = None
                fila[f"{prefix}_N_candidatos"] = None

        if error:
            print(f"  ERROR: {error}")

        resultados.append(fila)

    # ── Guardar ───────────────────────────────────────────────────────────────
    df          = pd.DataFrame(resultados)
    hoja_clean  = sheet_name.strip().replace(" ", "_")
    ks_suffix   = f"_k{k_min}" if k_min == k_max else f"_k{k_min}-{k_max}"
    eta_suffix  = f"_{etiqueta}" if etiqueta else ""
    ruta_salida = GEOMIP_ROOT / "results" / f"resultados_kgeomip_simetrico_{hoja_clean}{ks_suffix}{eta_suffix}.xlsx"
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(ruta_salida, index=False)

    completadas = sum(1 for r in resultados if r["Error"] is None)
    print(f"\n{'=' * 70}")
    print(f"Resultados guardados en: {ruta_salida}")
    print(f"Completadas: {completadas}/{len(resultados)}  |  Errores/Timeouts: {len(resultados)-completadas}")
    print("=" * 70)


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Ejecutor KGeoMIP Simetrico")
    parser.add_argument("--hoja",     required=True,       help="Nombre de hoja (ej. '10A-Elementos')")
    parser.add_argument("--inicio",   type=int, default=0, help="Indice de inicio 0-based (default 0)")
    parser.add_argument("--cantidad", type=int, default=50,help="Pruebas a ejecutar (default 50)")
    parser.add_argument("--timeout",  type=int, default=TIMEOUT_S,
                        help=f"Timeout por prueba en segundos (default {TIMEOUT_S})")
    parser.add_argument("--exacto", action="store_true",
                        help="Forzar enumeracion exhaustiva S(n,k) para todos los casos "
                             "(ignora n_max_exacto). Guarda con sufijo '_exacto'.")
    parser.add_argument("--etiqueta", type=str, default="",
                        help="Sufijo adicional para el nombre del archivo de salida")

    # Seleccion de k — mutuamente excluyentes: --k fija una sola k; --k_min/--k_max fijan rango
    k_group = parser.add_mutually_exclusive_group()
    k_group.add_argument("--k",     type=int,
                         help="Evaluar solo esta k (ej. --k 2 corre solo biparticiones)")
    k_group.add_argument("--k_min", type=int, default=None,
                         help="k minimo del rango (default 2, usar junto con --k_max)")
    parser.add_argument("--k_max",  type=int, default=None,
                        help="k maximo del rango (default 5, usar junto con --k_min)")

    args = parser.parse_args()

    if args.k is not None:
        k_min = k_max = args.k
    else:
        k_min = args.k_min if args.k_min is not None else 2
        k_max = args.k_max if args.k_max is not None else 5

    if k_min < 2:
        parser.error("k_min debe ser >= 2")
    if k_max > 5:
        parser.error("k_max debe ser <= 5")
    if k_min > k_max:
        parser.error("k_min no puede ser mayor que k_max")

    n_max_exacto = 999 if args.exacto else N_MAX_EXACTO
    etiqueta     = "exacto" if args.exacto else args.etiqueta

    ejecutar_desde_excel(
        sheet_name   = args.hoja,
        inicio       = args.inicio,
        cantidad     = args.cantidad,
        k_min        = k_min,
        k_max        = k_max,
        timeout_s    = args.timeout,
        n_max_exacto = n_max_exacto,
        etiqueta     = etiqueta,
    )


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()
