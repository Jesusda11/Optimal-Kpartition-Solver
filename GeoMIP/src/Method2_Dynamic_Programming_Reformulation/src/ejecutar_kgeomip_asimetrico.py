"""
src/ejecutar_kgeomip_asimetrico.py

Ejecutor de KGeoMIP Asimetrico para las pruebas del proyecto K-QGMIP.

Lee el Excel de pruebas (GeoMIP/tests/PruebasK-Particiones.xlsx), ejecuta
KGeoMIPAsimetrico y guarda resultados en:
    GeoMIP/results/resultados_kgeomip_asimetrico_{hoja}.xlsx

Los resultados de esta herramienta corresponden a las columnas "Geometric"
del Excel de pruebas (enfoque asimetrico), a diferencia de ejecutar_kgeomip.py
que genera las columnas "QNodes" (enfoque simetrico).

NO modifica el Excel de entrada. El usuario copia los resultados manualmente.

Como ejecutar (desde el directorio Method2_Dynamic_Programming_Reformulation/):

  Todas las k (2 a 5):
    python src/ejecutar_kgeomip_asimetrico.py --hoja 10A-Elementos

  Solo una k especifica:
    python src/ejecutar_kgeomip_asimetrico.py --hoja 10A-Elementos --k 2

  Rango de k:
    python src/ejecutar_kgeomip_asimetrico.py --hoja 10A-Elementos --k_min 2 --k_max 3

  Forzar exacto (util para n pequeno):
    python src/ejecutar_kgeomip_asimetrico.py --hoja 10A-Elementos --exacto

  Con paginacion:
    python src/ejecutar_kgeomip_asimetrico.py --hoja "25A-Elementos " --inicio 0 --cantidad 25

Variables de entorno opcionales:
    KGEOMIP_INPUT_XLSX      : ruta al Excel de entrada
    KGEOMIP_M_MAX_EXACTO    : umbral exacto/heuristico para pool m=n_fut+n_pres (default 8)
    KGEOMIP_TIMEOUT_S       : timeout por prueba en segundos (default 86400 = 24h)
"""

import sys
import os
import re
import time
import logging
import argparse
import multiprocessing
from math import comb
from pathlib import Path

import numpy as np
import pandas as pd

# ── Rutas ─────────────────────────────────────────────────────────────────────
METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT  = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(METHOD2_ROOT))

# ── Silenciar logging ─────────────────────────────────────────────────────────
logging.disable(logging.CRITICAL)

import src.middlewares.slogger as _slogger_mod

class _NullSafeLogger:
    def __init__(self, name): pass
    def debug(self, *a, **k): pass
    def info(self, *a, **k): pass
    def warn(self, *a, **k): pass
    def error(self, *a, **k): pass
    def critic(self, *a, **k): pass
    def fatal(self, *a, **k): pass
    def set_log(self, *a, **k): pass

_slogger_mod.SafeLogger = _NullSafeLogger

# ── Importaciones del proyecto ────────────────────────────────────────────────
from src.models.base.application import aplicacion
from src.middlewares.profile import profiler_manager
from src.controllers.manager import Manager
from src.controllers.strategies.kgeometric_asimetrico import KGeoMIPAsimetrico
from src.controllers.strategies.kgeometric_asimetrico_jerarquico import KGeoMIPAsimetricoJerarquico
from src.funcs.decay import decay_exponencial
from src.funcs.format import fmt_k_particion

profiler_manager.enabled = False

# ── Configuracion base ─────────────────────────────────────────────────────────
M_MAX_EXACTO      = int(os.getenv("KGEOMIP_M_MAX_EXACTO",      "8"))
M_MAX_CANDIDATOS  = int(os.getenv("KGEOMIP_M_MAX_CANDIDATOS", "2000"))
TIMEOUT_S         = int(os.getenv("KGEOMIP_TIMEOUT_S",        "86400"))  # 24h default
LIMITE_MEM_GB     = float(os.getenv("KGEOMIP_LIMITE_MEM_GB",  "4.0"))    # advertencia OOM
INPUT_XLSX   = Path(os.getenv(
    "KGEOMIP_INPUT_XLSX",
    str(GEOMIP_ROOT / "tests" / "PruebasK-Particiones.xlsx"),
))


# ── Helpers ───────────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int) -> str:
    posiciones = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto.upper():
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def parsear_hoja(sheet_name: str) -> tuple[int, str]:
    m = re.match(r"(\d+)([A-Z])", sheet_name.strip())
    if not m:
        raise ValueError(f"No se pudo parsear '{sheet_name}'. Formato: '10A-Elementos'.")
    return int(m.group(1)), m.group(2)


def resolver_tpm(n: int, variante: str) -> tuple[Path, np.ndarray]:
    nombre = f"N{n}{variante}.csv"
    for base in (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT  / "data" / "samples",
    ):
        if (base / nombre).exists():
            p = base / nombre
            return p, np.genfromtxt(p, delimiter=",")
    raise FileNotFoundError(f"No se encontro '{nombre}'")


def leer_excel(ruta: Path, sheet_name: str) -> tuple[str, list[tuple]]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[sheet_name]
    raw = ws["B1"].value
    estado_inicial = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    filas = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and alcance_str is not None and mecanismo_str is not None:
            filas.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))
    wb.close()
    return estado_inicial, filas


def _formatear_particion_asimetrica(grupos: list) -> str | None:
    """
    Formatea la particion asimetrica optima para mostrar en el Excel.

    A diferencia del caso simetrico, futuros y presentes son listas independientes
    por grupo, asi que se pasan directamente a fmt_k_particion.
    Grupos con futuros o presentes vacios se incluyen tal cual (el formato
    mostrara el simbolo de vacio para el lado correspondiente).
    """
    if not grupos:
        return None
    grupos_fmt = [(list(map(int, f)), list(map(int, p))) for f, p in grupos]
    return fmt_k_particion(grupos_fmt)


# ── Worker de multiprocessing ─────────────────────────────────────────────────

def _estimar_mem_gb(n_pres: int, m: int, k_max: int) -> float:
    """Estima RAM necesaria para el modo directo (GB). ~700 bytes por frozenset."""
    if m < 2 or k_max < 2:
        return 0.0
    bfs_estados = sum(comb(n_pres, i) for i in range(1, n_pres // 2 + 1))
    cuts_por_estado = comb(m - 1, k_max - 1)
    return bfs_estados * cuts_por_estado * 700 / 1e9


def _worker(
    estado_inicial:   str,
    condiciones:      str,
    alcance:          str,
    mecanismo:        str,
    tpm:              "np.ndarray",
    pagina:           str,
    m_max_exacto:     int,
    k_min:            int,
    k_max:            int,
    cola:             multiprocessing.Queue,
    m_max_candidatos: int = 2000,
    modo:             str = "directo",
) -> None:
    """
    Proceso hijo: crea la estrategia asimetrica, ejecuta y pone resultados en la cola.

    modo="directo"     -> KGeoMIPAsimetrico (generacion directa de candidatos)
    modo="jerarquico"  -> KGeoMIPAsimetricoJerarquico (biparticiones sucesivas)
    """
    try:
        aplicacion.pagina_sample_network = pagina
        gestor = Manager(estado_inicial=estado_inicial)

        Clase = KGeoMIPAsimetricoJerarquico if modo == "jerarquico" else KGeoMIPAsimetrico
        sia = Clase(
            gestor,
            k_max=k_max,
            k_min=k_min,
            m_max_exhaustivo=m_max_exacto,
            m_max_candidatos=m_max_candidatos,
            decay_fn=decay_exponencial,
        )
        sia.aplicar_estrategia(condiciones, alcance, mecanismo, tpm)

        resultado = {"modo": sia._modo_usado, "por_k": {}, "error": None}

        for k, datos in sia._resultados_por_k.items():
            grupos = datos.get("grupos") or []
            particion_str = _formatear_particion_asimetrica(grupos)
            resultado["por_k"][k] = {
                "phi":          datos["phi"],
                "particion":    particion_str,
                "n_candidatos": datos["n_candidatos"],
                "tiempo_s":     datos["tiempo_s"],
                "modo_usado":   datos.get("modo_usado", sia._modo_usado),
            }

    except Exception as exc:
        resultado = {"modo": "error", "por_k": {}, "error": str(exc)}

    cola.put(resultado)


# ── Ejecutor principal ────────────────────────────────────────────────────────

def ejecutar_desde_excel(
    sheet_name:       str,
    inicio:           int = 0,
    cantidad:         int = 50,
    k_min:            int = 2,
    k_max:            int = 5,
    timeout_s:        int = TIMEOUT_S,
    m_max_exacto:     int = M_MAX_EXACTO,
    m_max_candidatos: int = M_MAX_CANDIDATOS,
    etiqueta:         str = "",
    modo:             str = "directo",
) -> None:
    n, variante = parsear_hoja(sheet_name)
    estado_inicial, todos_casos = leer_excel(INPUT_XLSX, sheet_name)
    tpm_path, tpm = resolver_tpm(n, variante)
    condiciones   = "1" * len(estado_inicial)
    filas         = todos_casos[inicio : inicio + cantidad]

    m_sistema = 2 * n  # pool size para sistema completamente balanceado (cota superior)
    ks_label  = f"k={k_min}" if k_min == k_max else f"k={k_min}..{k_max}"
    modo_label = "EXACTO EXHAUSTIVO" if m_max_exacto >= 999 else f"m_max_exacto={m_max_exacto}"

    print("=" * 70)
    print(f"KGeoMIP Asimetrico — Hoja: {sheet_name}  |  {ks_label}  |  {modo_label}")
    print(f"Modo de busqueda: {modo.upper()}")
    print(f"Estado inicial: {estado_inicial}  n={len(estado_inicial)}  m_sistema_max={m_sistema}")
    print(f"TPM: {tpm_path}")
    print(f"Pruebas: {inicio + 1} -> {inicio + len(filas)}")
    print(f"Candidatos heuristicos max: {m_max_candidatos}")
    print(f"Timeout por prueba: {timeout_s}s ({timeout_s/3600:.1f}h)")
    if m_max_exacto >= 999:
        from src.funcs.partitions import contar_stirling
        stirling_max = sum(contar_stirling(m_sistema, k) for k in range(k_min, k_max + 1))
        print(f"AVISO: modo exacto — hasta {stirling_max:,} candidatos por caso (m_max={m_sistema})")
    print("=" * 70)

    resultados = []

    for prueba_num, alcance_str, mecanismo_str in filas:
        alcance   = convertir_a_binario(alcance_str,   len(estado_inicial))
        mecanismo = convertir_a_binario(mecanismo_str, len(estado_inicial))

        print(f"\n[{prueba_num:>3}] Alcance={alcance_str!r:<30} Mecanismo={mecanismo_str!r}")

        # Advertencia de memoria para modo directo
        if modo == "directo" and m_max_exacto < 999:
            n_fut_est  = sum(b == "1" and c == "1" for b, c in zip(alcance,   condiciones))
            n_pres_est = sum(b == "1" and c == "1" for b, c in zip(mecanismo, condiciones))
            m_est      = n_fut_est + n_pres_est
            if m_est > m_max_exacto:
                mem_gb = _estimar_mem_gb(n_pres_est, m_est, k_max)
                if mem_gb > LIMITE_MEM_GB:
                    print(
                        f"  ADVERTENCIA: modo directo estima ~{mem_gb:.1f} GB para "
                        f"k={k_max}, m={m_est} (n_pres={n_pres_est}). "
                        f"Usar --modo jerarquico para evitar OOM."
                    )

        cola    = multiprocessing.Queue()
        proceso = multiprocessing.Process(
            target=_worker,
            args=(estado_inicial, condiciones, alcance, mecanismo,
                  tpm, variante, m_max_exacto, k_min, k_max, cola,
                  m_max_candidatos, modo),
        )
        t_ini = time.perf_counter()
        proceso.start()
        proceso.join(timeout=timeout_s)
        t_total = round(time.perf_counter() - t_ini, 4)

        if proceso.is_alive():
            print(f"  [TIMEOUT {timeout_s}s] terminado forzosamente.")
            proceso.terminate()
            proceso.join()
            por_k_data, modo, error = {}, "timeout", f"Timeout ({timeout_s}s)"
        elif cola.empty():
            print(f"  [DEBUG] exitcode={proceso.exitcode}")
            por_k_data, modo, error = {}, "error", "Proceso termino sin resultado"
        else:
            res        = cola.get()
            por_k_data = res.get("por_k", {})
            modo       = res.get("modo", "?")
            error      = res.get("error")

        fila: dict = {
            "#Prueba":   prueba_num,
            "Alcance":   alcance_str,
            "Mecanismo": mecanismo_str,
            "Modo":      modo,
            "T_total_s": t_total,
            "Error":     error,
        }

        for k in range(2, 6):  # columnas k2-k5 para compatibilidad con Excel
            prefix = f"k{k}"
            if k in por_k_data:
                d = por_k_data[k]
                phi_str = str(d["phi"]).replace(".", ",") if d["phi"] is not None else None
                t_str   = str(d["tiempo_s"]).replace(".", ",") if d["tiempo_s"] is not None else None
                fila[f"{prefix}_Particion"]    = d["particion"]
                fila[f"{prefix}_Perdida"]      = phi_str
                fila[f"{prefix}_Tiempo_s"]     = t_str
                fila[f"{prefix}_N_candidatos"] = d["n_candidatos"]
                fila[f"{prefix}_ModoUsado"]    = d.get("modo_usado", modo)
                if error is None:
                    print(
                        f"  k={k}: phi={d['phi']:.6f}  t={d['tiempo_s']:.3f}s"
                        f"  cand={d['n_candidatos']}  modo={d.get('modo_usado', modo)}"
                    )
            else:
                fila[f"{prefix}_Particion"]    = None
                fila[f"{prefix}_Perdida"]      = None
                fila[f"{prefix}_Tiempo_s"]     = None
                fila[f"{prefix}_N_candidatos"] = None
                fila[f"{prefix}_ModoUsado"]    = None

        if error:
            print(f"  ERROR: {error}")

        resultados.append(fila)

    # ── Guardar ───────────────────────────────────────────────────────────────
    df         = pd.DataFrame(resultados)
    hoja_clean = sheet_name.strip().replace(" ", "_")
    ks_suffix  = f"_k{k_min}" if k_min == k_max else f"_k{k_min}-{k_max}"
    # Solo agrega sufijo de rango cuando NO es la corrida completa por defecto (inicio=0, cantidad=50).
    # Asi el archivo sin sufijo de rango es exclusivo para las 50 pruebas completas.
    rango_suffix = f"_p{inicio + 1}-{inicio + cantidad}" if (inicio != 0 or cantidad != 50) else ""
    eta_suffix = f"_{etiqueta}" if etiqueta else ""
    ruta_salida = (
        GEOMIP_ROOT / "results"
        / f"resultados_kgeomip_asimetrico_{hoja_clean}{ks_suffix}{rango_suffix}{eta_suffix}.xlsx"
    )
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(ruta_salida, index=False)

    completadas = sum(1 for r in resultados if r["Error"] is None)
    print(f"\n{'=' * 70}")
    print(f"Resultados guardados en: {ruta_salida}")
    print(f"Completadas: {completadas}/{len(resultados)}  |  Errores/Timeouts: {len(resultados)-completadas}")
    print("=" * 70)


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Ejecutor KGeoMIP Asimetrico")
    parser.add_argument("--hoja",     required=True,        help="Nombre de hoja (ej. '10A-Elementos')")
    parser.add_argument("--inicio",   type=int, default=0,  help="Indice de inicio 0-based (default 0)")
    parser.add_argument("--cantidad", type=int, default=50, help="Pruebas a ejecutar (default 50)")
    parser.add_argument("--timeout",  type=int, default=TIMEOUT_S,
                        help=f"Timeout por prueba en segundos (default {TIMEOUT_S})")
    parser.add_argument("--exacto", action="store_true",
                        help="Forzar enumeracion exhaustiva S(m,k) para todos los casos.")
    parser.add_argument("--m_max_candidatos", type=int, default=M_MAX_CANDIDATOS,
                        help=f"Limite de candidatos heuristicos por k (default {M_MAX_CANDIDATOS}). "
                             "Reducir (ej. 500) alivia presion de RAM en sistemas grandes (n>=20).")
    parser.add_argument("--etiqueta", type=str, default="",
                        help="Sufijo adicional para el nombre del archivo de salida")
    parser.add_argument("--modo", choices=["directo", "jerarquico"], default="directo",
                        help="Modo de busqueda: directo (default, exhaustivo S(m,k)) o "
                             "jerarquico (biparticiones sucesivas, recomendado para n>=15, k>=4).")

    k_group = parser.add_mutually_exclusive_group()
    k_group.add_argument("--k",     type=int,
                         help="Evaluar solo esta k (ej. --k 2)")
    k_group.add_argument("--k_min", type=int, default=None,
                         help="k minimo del rango (default 2)")
    parser.add_argument("--k_max",  type=int, default=None,
                        help="k maximo del rango (default 5)")

    args = parser.parse_args()

    if args.k is not None:
        k_min = k_max = args.k
    else:
        k_min = args.k_min if args.k_min is not None else 2
        k_max = args.k_max if args.k_max is not None else 5

    if k_min < 2:
        parser.error("k_min debe ser >= 2")
    if k_max > 5:
        parser.error("k_max debe ser <= 5")
    if k_min > k_max:
        parser.error("k_min no puede ser mayor que k_max")

    m_max_exacto = 999 if args.exacto else M_MAX_EXACTO
    etiqueta     = "exacto" if args.exacto else args.etiqueta
    modo         = args.modo

    ejecutar_desde_excel(
        sheet_name       = args.hoja,
        inicio           = args.inicio,
        cantidad         = args.cantidad,
        k_min            = k_min,
        k_max            = k_max,
        timeout_s        = args.timeout,
        m_max_exacto     = m_max_exacto,
        m_max_candidatos = args.m_max_candidatos,
        etiqueta         = etiqueta,
        modo             = modo,
    )


if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()
