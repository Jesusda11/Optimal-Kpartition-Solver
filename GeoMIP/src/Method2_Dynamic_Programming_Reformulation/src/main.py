from src.controllers.manager import Manager
from src.controllers.strategies.geometric import GeometricSIA
from src.controllers.strategies.q_nodes import QNodes

try:
    from src.controllers.strategies.phi import Phi
except Exception:
    Phi = None

import multiprocessing
import numpy as np
import pandas as pd
import os
import re
from pathlib import Path


METHOD2_ROOT = Path(__file__).resolve().parents[1]
GEOMIP_ROOT  = Path(__file__).resolve().parents[3]


# ─────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int = 20) -> str:
    """Convert a letter-string like 'ABCDFG' to a binary string of length n_bits."""
    posiciones = "ABCDEFGHIJKLMNOPQRST"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto:
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def ejecutar_con_tiempo(config_sistema, condiciones, alcance, mecanismo, resultado_queue, tpm):
    try:
        analizador_fi = GeometricSIA(config_sistema)
        sia_dos = analizador_fi.aplicar_estrategia(condiciones, alcance, mecanismo, tpm)
        resultado_queue.put({
            "particion": sia_dos.particion,
            "perdida":   str(sia_dos.perdida).replace(".", ","),
            "tiempo":    str(sia_dos.tiempo_ejecucion).replace(".", ","),
        })
    except Exception as e:
        print(f"[ERROR en proceso hijo] {e}")
        resultado_queue.put({"particion": None, "perdida": None, "tiempo": None})


def resolver_tpm_path(estado_inicio: str) -> Path:
    """Find TPM file in common project locations based on state size."""
    sample_name = f"N{len(estado_inicio)}A.csv"
    candidates = (
        METHOD2_ROOT / "src" / ".samples" / sample_name,
        METHOD2_ROOT / ".samples"          / sample_name,
        GEOMIP_ROOT  / "data" / "samples"  / sample_name,
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"No se encontró la TPM '{sample_name}'. "
        f"Busqué en: {', '.join(str(c) for c in candidates)}"
    )


def inferir_estado_inicial() -> str:
    """Infer an initial state from available datasets (prefers largest NxA.csv)."""
    sample_dirs = (
        METHOD2_ROOT / "src" / ".samples",
        METHOD2_ROOT / ".samples",
        GEOMIP_ROOT  / "data" / "samples",
    )
    pattern = re.compile(r"N(\d+)[A-Z]\.csv$")
    available_sizes = []

    for sample_dir in sample_dirs:
        if not sample_dir.exists():
            continue
        for sample_file in sample_dir.glob("N*.csv"):
            match = pattern.match(sample_file.name)
            if match:
                available_sizes.append(int(match.group(1)))

    if not available_sizes:
        raise FileNotFoundError(
            "No hay archivos de muestras TPM disponibles en data/samples ni .samples."
        )

    n_bits = max(available_sizes)
    return "1" + ("0" * (n_bits - 1))


# ─────────────────────────────────────────────────────────────
#  Core runner — adapted for the new Excel layout
#
#  New layout (DatosPruebas2026_1.xlsx, sheet "10A-Elementos"):
#    Row 1 : Estado inicial  → col B  (e.g. 1000000000)
#    Row 2 : Sistema         → col B
#    Row 3 : Sistema cand.   → col B
#    Row 4 : section label   → col B
#    Row 5 : column headers  → #Prueba | Alcance (B) | Mecanismo (C) | …
#    Row 6+: data rows       → int    | letter-string | letter-string
# ─────────────────────────────────────────────────────────────

def ejecutar_desde_excel(
    ruta_excel:   Path,
    ruta_salida:  Path,
    sheet_name:   str        = "10A-Elementos",
    inicio:       int        = 0,
    cantidad:     int        = 50,
    estado_inicio: str | None = None,
    condiciones:  str | None  = None,
):
    # ── 1. Read metadata from header rows ──────────────────────────────────
    from openpyxl import load_workbook
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Row 1, col B  →  estado inicial (may come back as int, e.g. 1000000000)
    raw_estado = ws["B1"].value
    if estado_inicio is None:
        estado_inicio = str(int(raw_estado)) if raw_estado is not None else inferir_estado_inicial()

    n_bits      = len(estado_inicio)
    condiciones = condiciones or ("1" * n_bits)
    tpm_path    = resolver_tpm_path(estado_inicio)
    tpm         = np.genfromtxt(tpm_path, delimiter=",")

    print(f"[Config] estado_inicio={estado_inicio}  n_bits={n_bits}")
    print(f"[Config] TPM cargada desde: {tpm_path}")
    print(f"[Config] Hoja: '{sheet_name}'  |  Pruebas: {inicio+1} → {inicio+cantidad}")

    # ── 2. Read data rows (start at Excel row 6 = openpyxl min_row=6) ──────
    all_rows = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and alcance_str is not None and mecanismo_str is not None:
            all_rows.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))

    wb.close()

    filas = all_rows[inicio : inicio + cantidad]
    resultados = []

    # ── 3. Run each test ────────────────────────────────────────────────────
    for prueba_num, alcance_str, mecanismo_str in filas:
        alcance   = convertir_a_binario(alcance_str,   n_bits=n_bits)
        mecanismo = convertir_a_binario(mecanismo_str, n_bits=n_bits)

        print(f"\nIteración {prueba_num} — Alcance: {alcance_str!r} → {alcance}")
        print(f"              Mecanismo: {mecanismo_str!r} → {mecanismo}")

        config_sistema  = Manager(estado_inicial=estado_inicio)
        resultado_queue = multiprocessing.Queue()

        proceso = multiprocessing.Process(
            target=ejecutar_con_tiempo,
            args=(config_sistema, condiciones, alcance, mecanismo, resultado_queue, tpm),
        )
        proceso.start()
        proceso.join(timeout=3600)   # 1-hour hard limit per test

        if proceso.is_alive():
            print(f"Iteración {prueba_num} — Tiempo límite alcanzado, terminando proceso…")
            proceso.terminate()
            proceso.join()
            resultado = {"perdida": None, "tiempo": None, "particion": None}
        else:
            resultado = (
                resultado_queue.get()
                if not resultado_queue.empty()
                else {"perdida": None, "tiempo": None, "particion": None}
            )

        resultados.append({
            "Iteración":              prueba_num,
            "Alcance":                alcance,
            "Mecanismo":              mecanismo,
            "Partición":              resultado["particion"],
            "Pérdida":                resultado["perdida"],
            "Tiempo de ejecución (s)": resultado["tiempo"],
        })

    # ── 4. Save results ─────────────────────────────────────────────────────
    df_resultados = pd.DataFrame(resultados)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    df_resultados.to_excel(ruta_salida, index=False)
    print(f"\n[OK] Resultados guardados en {ruta_salida}")


# ─────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────

def iniciar():
    ruta_entrada = Path(
        os.getenv(
            "GEOMIP_INPUT_XLSX",
            str(GEOMIP_ROOT / "tests" / "DatosPruebas2026_1.xlsx"),
        )
    )
    ruta_salida = Path(
        os.getenv(
            "GEOMIP_OUTPUT_XLSX",
            str(GEOMIP_ROOT / "results" / "resultados_Geometric_15B.xlsx"),
        )
    )

    ejecutar_desde_excel(
        ruta_excel    = ruta_entrada,
        ruta_salida   = ruta_salida,
        sheet_name    = "15B-Elementos",   # hoja de 15 variables
        inicio        = 0,
        cantidad      = 50,
        estado_inicio = "100000000000000", # estado inicial fijo para N=15
    )