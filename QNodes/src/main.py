from src.controllers.manager import Manager
from src.strategies.q_nodes import QNodes

import multiprocessing
import pandas as pd
import os
from pathlib import Path


QNODES_ROOT = Path(__file__).resolve().parents[1]


# ─────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────

def convertir_a_binario(texto: str, n_bits: int) -> str:
    """Convierte una cadena de letras como 'ABCDFG' a binario de longitud n_bits."""
    posiciones = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:n_bits]
    binario = ["0"] * n_bits
    for letra in texto:
        if letra in posiciones:
            binario[posiciones.index(letra)] = "1"
    return "".join(binario)


def ejecutar_con_tiempo(tpm, estado_inicio, condiciones, alcance, mecanismo, resultado_queue):
    try:
        analizador_q = QNodes(tpm)
        resultado = analizador_q.aplicar_estrategia(
            estado_inicio,
            condiciones,
            alcance,
            mecanismo,
        )
        resultado_queue.put({
            "particion": resultado.particion,
            "perdida":   str(resultado.perdida).replace(".", ","),
            "tiempo":    str(resultado.tiempo_ejecucion).replace(".", ","),
        })
    except Exception as e:
        print(f"[ERROR en proceso hijo] {e}")
        resultado_queue.put({"particion": None, "perdida": None, "tiempo": None})


# ─────────────────────────────────────────────────────────────
#  Core runner
#
#  Layout Excel (DatosPruebas2026_1.xlsx):
#    Fila 1 : Estado inicial  → col B  (e.g. 1000000000)
#    Fila 5 : Headers         → #Prueba | Alcance (B) | Mecanismo (C)
#    Fila 6+: Datos           → int    | letras        | letras
# ─────────────────────────────────────────────────────────────

def ejecutar_desde_excel(
    ruta_excel:    Path,
    ruta_salida:   Path,
    sheet_name:    str       = "20A-Elementos",
    inicio:        int       = 0,
    cantidad:      int       = 50,
    estado_inicio: str | None = None,
    condiciones:   str | None = None,
):
    # ── 1. Leer metadata y datos del Excel ─────────────────────────────────
    from openpyxl import load_workbook
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb[sheet_name]

    raw_estado = ws["B1"].value
    if estado_inicio is None:
        estado_inicio = str(int(raw_estado)) if raw_estado is not None else "1000000000"

    n_bits      = len(estado_inicio)
    condiciones = condiciones or ("1" * n_bits)

    all_rows = []
    for row in ws.iter_rows(min_row=6, max_col=3, values_only=True):
        prueba, alcance_str, mecanismo_str = row
        if prueba is not None and alcance_str is not None and mecanismo_str is not None:
            all_rows.append((int(prueba), str(alcance_str).strip(), str(mecanismo_str).strip()))

    wb.close()

    filas = all_rows[inicio : inicio + cantidad]

    # ── 2. Cargar TPM una sola vez via Manager ─────────────────────────────
    gestor = Manager(estado_inicio)
    tpm    = gestor.cargar_red()

    print(f"[Config] estado_inicio={estado_inicio}  n_bits={n_bits}")
    print(f"[Config] Hoja: '{sheet_name}'  |  Pruebas: {inicio+1} → {inicio+cantidad}")

    # ── 3. Ejecutar cada prueba ────────────────────────────────────────────
    resultados = []

    for prueba_num, alcance_str, mecanismo_str in filas:
        alcance   = convertir_a_binario(alcance_str,   n_bits=n_bits)
        mecanismo = convertir_a_binario(mecanismo_str, n_bits=n_bits)

        print(f"\nIteración {prueba_num} — Alcance: {alcance_str!r} → {alcance}")
        print(f"               Mecanismo: {mecanismo_str!r} → {mecanismo}")

        resultado_queue = multiprocessing.Queue()
        proceso = multiprocessing.Process(
            target=ejecutar_con_tiempo,
            args=(tpm, estado_inicio, condiciones, alcance, mecanismo, resultado_queue),
        )

        proceso.start()
        proceso.join(timeout=3600)  # límite de 1 hora por prueba

        if proceso.is_alive():
            print(f"Iteración {prueba_num} — Tiempo límite alcanzado, terminando proceso...")
            proceso.terminate()
            proceso.join()
            resultado = {"particion": None, "perdida": None, "tiempo": None}
        else:
            resultado = (
                resultado_queue.get()
                if not resultado_queue.empty()
                else {"particion": None, "perdida": None, "tiempo": None}
            )

        print(f"  → Partición : {resultado['particion']}")
        print(f"  → Pérdida   : {resultado['perdida']}")
        print(f"  → Tiempo (s): {resultado['tiempo']}")

        resultados.append({
            "Iteración":               prueba_num,
            "Alcance":                 alcance,
            "Mecanismo":               mecanismo,
            "Partición":               resultado["particion"],
            "Pérdida":                 resultado["perdida"],
            "Tiempo de ejecución (s)": resultado["tiempo"],
        })

    # ── 4. Guardar resultados ──────────────────────────────────────────────
    df_resultados = pd.DataFrame(resultados)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    df_resultados.to_excel(ruta_salida, index=False)
    print(f"\n[OK] Resultados guardados en {ruta_salida}")


# ─────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────

def iniciar():
    ruta_entrada = Path(
        os.getenv(
            "QNODES_INPUT_XLSX",
            str(QNODES_ROOT / "tests" / "DatosPruebas2026_1.xlsx"),
        )
    )
    ruta_salida = Path(
        os.getenv(
            "QNODES_OUTPUT_XLSX",
            str(QNODES_ROOT / "results" / "resultados_QNodes_20A.xlsx"),
        )
    )

    ejecutar_desde_excel(
        ruta_excel    = ruta_entrada,
        ruta_salida   = ruta_salida,
        sheet_name    = "20A-Elementos",
        inicio        = 0,
        cantidad      = 50,
        estado_inicio = "10000000000000000000",   # N=20
    )